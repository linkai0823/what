# ！/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
© Copyright 2018 The Author. All Rights Reversed.
------------------------------------------------------------
File Name:
Author : zhangtao
Time:
Description：
------------------------------------------------------------
"""

import pymysql.cursors
import regex_openlaw as openlaw
import openpyxl

connection = pymysql.connect(host='localhost',
                             user='root',
                             password='1234qwer',
                             db='test',
                             charset='utf8mb4',
                             cursorclass=pymysql.cursors.DictCursor)
try:

    wb=openpyxl.load_workbook('OpenLaw_cases.xlsx')  #打开excel文件

    sheet=wb.worksheets[0]  #获取工作表
    print(sheet.title)

    for r in range(2, sheet.max_row + 1):
        # for r in range(2, 100):
        title = sheet.cell(row=r, column=1).value
        case_number = sheet.cell(row=r, column=2).value

        with connection.cursor() as cursor:
            sql = "SELECT openlawdata.`案号` FROM openlawdata WHERE `案号` = '%s'" % (case_number)
            cursor.execute(sql)
            results = cursor.fetchall()
        connection.commit()
        print(results)

        # case_type = sheet.cell(row=r, column=3).value
        # procedure_of_trial = sheet.cell(row=r, column=4).value
        # cause_of_action = sheet.cell(row=r, column=5).value
        # document_type = sheet.cell(row=r, column=6).value
        # court = sheet.cell(row=r, column=7).value
        # judgment_date = sheet.cell(row=r, column=8).value
        # plaintiff = openlaw.split_text(sheet.cell(row=r, column=9).value)
        # defendant = openlaw.split_text(sheet.cell(row=r, column=10).value)
        # judge = sheet.cell(row=r, column=12).value
        # presiding_judge = sheet.cell(row=r, column=13).value
        # judicial_officer = sheet.cell(row=r, column=14).value
        # court_clerk = sheet.cell(row=r, column=15).value
        #
        # print('defendant:', defendant)
        #
        #
        # if defendant == ['']:
        #     pass
        # else:
        #     print(len(defendant))
        #     for i in range(len(defendant)):
        #
        #         with connection.cursor() as cursor:
        #             # Create a new record
        #             sql = "INSERT IGNORE INTO `defendant` (`case_number`, `defendant`) VALUES (%s, %s)"
        #             cursor.execute(sql, (case_number, defendant[i]))
        #         connection.commit()
        #


finally:
    connection.close()

